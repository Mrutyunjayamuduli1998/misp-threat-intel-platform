import openpyxl
import requests
import json
import urllib3
from datetime import datetime

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Configuration ──
MISP_URL = "https://localhost"
MISP_KEY = input("Enter your MISP API key: ").strip()
EXCEL_FILE = "/mnt/c/Users/arpit/Documents/ioc_intel.xlsx"
VERIFY_SSL = False

headers = {
    "Authorization": MISP_KEY,
    "Content-Type": "application/json",
    "Accept": "application/json"
}

# ── Confidence → MISP Threat Level mapping ──
CONFIDENCE_MAP = {
    "low":      {"threat_level_id": "4", "label": "Low"},
    "medium":   {"threat_level_id": "3", "label": "Medium"},
    "high":     {"threat_level_id": "2", "label": "High"},
    "critical": {"threat_level_id": "1", "label": "Critical"},
}

# ── Threat Type → MISP Tag mapping ──
THREAT_TYPE_TAGS = {
    "Phishing":                  "misp-galaxy:mitre-attack-pattern=\"Phishing\"",
    "Ransomware":                "misp-galaxy:ransomware",
    "APT/Espionage":             "misp-galaxy:threat-actor",
    "C2 Infrastructure":         "misp-galaxy:mitre-attack-pattern=\"Command and Control\"",
    "Malware Distribution":      "misp-galaxy:mitre-attack-pattern=\"Ingress Tool Transfer\"",
    "DDoS":                      "misp-galaxy:mitre-attack-pattern=\"Network Denial of Service\"",
    "Supply Chain":              "misp-galaxy:mitre-attack-pattern=\"Supply Chain Compromise\"",
    "Vulnerability Exploitation":"misp-galaxy:mitre-attack-pattern=\"Exploit Public-Facing Application\"",
}

def check_existing_ioc(value):
    """Check if IOC already exists in MISP"""
    try:
        resp = requests.post(
            f"{MISP_URL}/attributes/restSearch",
            headers=headers,
            json={"value": value, "limit": 1},
            verify=VERIFY_SSL,
            timeout=10
        )
        data = resp.json()
        attrs = data.get("response", {}).get("Attribute", [])
        if attrs:
            return attrs[0].get("event_id")
        return None
    except:
        return None

def create_event(row):
    """Create a new MISP event from Excel row"""
    threat_level = CONFIDENCE_MAP.get(
        str(row.get("confidence", "medium")).lower(),
        CONFIDENCE_MAP["medium"]
    )

    # Build event info string
    parts = []
    if row.get("threat_actor"):
        parts.append(row["threat_actor"])
    if row.get("threat_type"):
        parts.append(row["threat_type"])
    if row.get("targeted_sector"):
        parts.append(f"targeting {row['targeted_sector']}")
    info = " | ".join(parts) if parts else "IOC Intelligence"

    event = {
        "Event": {
            "info": info,
            "threat_level_id": threat_level["threat_level_id"],
            "analysis": "2",
            "distribution": "0",
            "date": str(datetime.now().date()),
        }
    }

    resp = requests.post(
        f"{MISP_URL}/events/add",
        headers=headers,
        json=event,
        verify=VERIFY_SSL,
        timeout=10
    )
    result = resp.json()
    return result.get("Event", {}).get("id")

def add_attribute(event_id, row):
    """Add IOC attribute to event"""
    attr = {
        "Attribute": {
            "event_id": event_id,
            "type": row.get("ioc_type", "domain"),
            "category": "Artifacts dropped",
            "value": row["indicator"],
            "to_ids": True,
            "comment": row.get("notes", ""),
            "first_seen": row.get("first_seen", ""),
            "last_seen": row.get("last_seen", ""),
        }
    }

    resp = requests.post(
        f"{MISP_URL}/attributes/add/{event_id}",
        headers=headers,
        json=attr,
        verify=VERIFY_SSL,
        timeout=10
    )
    result = resp.json()
    return result.get("Attribute", {}).get("id")

def add_tags(event_id, row):
    """Add all relevant tags to event"""
    tags = []

    # TLP tag based on confidence
    conf = str(row.get("confidence", "")).lower()
    if conf == "critical":
        tags.append("tlp:red")
    elif conf == "high":
        tags.append("tlp:amber")
    elif conf == "medium":
        tags.append("tlp:green")
    else:
        tags.append("tlp:white")

    # Threat type tag
    threat_type = row.get("threat_type", "")
    if threat_type in THREAT_TYPE_TAGS:
        tags.append(THREAT_TYPE_TAGS[threat_type])

    # Threat actor tag
    if row.get("threat_actor"):
        tags.append(f"threat-actor:{row['threat_actor']}")

    # Country tags
    if row.get("country_origin"):
        tags.append(f"country-of-origin:{row['country_origin']}")
    if row.get("targeted_country"):
        tags.append(f"targeted-country:{row['targeted_country']}")
    if row.get("targeted_region"):
        tags.append(f"targeted-region:{row['targeted_region']}")
    if row.get("targeted_sector"):
        tags.append(f"targeted-sector:{row['targeted_sector']}")

    # Malware tag
    if row.get("associated_malware"):
        tags.append(f"malware:{row['associated_malware']}")

    # CVE tag
    if row.get("associated_cve"):
        tags.append(f"vulnerability:{row['associated_cve']}")

    # Custom tags from Excel
    if row.get("tags"):
        for t in str(row["tags"]).split(","):
            tags.append(t.strip())

    # Add each tag to event
    for tag in tags:
        try:
            requests.post(
                f"{MISP_URL}/events/addTag/{event_id}/{requests.utils.quote(tag)}",
                headers=headers,
                verify=VERIFY_SSL,
                timeout=5
            )
        except:
            pass

def add_objects(event_id, row):
    """Add extra context as MISP attributes"""
    extra_attrs = []

    if row.get("reference_url"):
        extra_attrs.append({
            "type": "url",
            "category": "External analysis",
            "value": row["reference_url"],
            "comment": "Source reference"
        })

    if row.get("associated_cve"):
        extra_attrs.append({
            "type": "vulnerability",
            "category": "External analysis",
            "value": row["associated_cve"],
            "comment": f"Associated CVE for {row['indicator']}"
        })

    if row.get("associated_malware"):
        extra_attrs.append({
            "type": "md5",
            "category": "Artifacts dropped",
            "value": row["associated_malware"],
            "comment": f"Associated malware family"
        })

    for attr in extra_attrs:
        try:
            attr["event_id"] = event_id
            requests.post(
                f"{MISP_URL}/attributes/add/{event_id}",
                headers=headers,
                json={"Attribute": attr},
                verify=VERIFY_SSL,
                timeout=5
            )
        except:
            pass

def main():
    print("\n" + "="*60)
    print("  MISP IOC Ingestion Tool")
    print("  Reading from: ioc_intel.xlsx")
    print("="*60 + "\n")

    # Load Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["IOC Intelligence"]

    # Get headers
    headers_row = [cell.value.lower() if cell.value else "" for cell in ws[1]]

    success = 0
    skipped = 0
    updated = 0
    errors = 0

    # Process each row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue

        # Map row to dict
        row_data = {headers_row[i]: str(v).strip() if v else "" 
                   for i, v in enumerate(row)}

        indicator = row_data.get("indicator", "").strip()
        if not indicator or indicator == "None":
            continue

        print(f"Processing [{row_idx-1}]: {indicator[:50]}")

        # Check for duplicate
        existing_event_id = check_existing_ioc(indicator)

        if existing_event_id:
            print(f"  → Found existing event {existing_event_id}, updating...")
            add_tags(existing_event_id, row_data)
            updated += 1
        else:
            # Create new event
            event_id = create_event(row_data)
            if not event_id:
                print(f"  → ERROR creating event")
                errors += 1
                continue

            # Add main IOC attribute
            attr_id = add_attribute(event_id, row_data)
            if not attr_id:
                print(f"  → ERROR adding attribute")
                errors += 1
                continue

            # Add tags
            add_tags(event_id, row_data)

            # Add extra context
            add_objects(event_id, row_data)

            print(f"  → Created event {event_id} ✓")
            success += 1

    print("\n" + "="*60)
    print(f"  ✅ Created:  {success} new events")
    print(f"  🔄 Updated:  {updated} existing events")
    print(f"  ❌ Errors:   {errors}")
    print(f"  Total processed: {success + updated + errors}")
    print("="*60)
    print(f"\nVerify in MISP: https://localhost/events/index\n")

if __name__ == "__main__":
    main()
# Category map fix - run this separately
CATEGORY_MAP = {
    "ip-dst": "Network activity",
    "ip-src": "Network activity", 
    "domain": "Network activity",
    "url": "Network activity",
    "md5": "Artifacts dropped",
    "sha256": "Artifacts dropped",
    "sha1": "Artifacts dropped",
    "email-src": "Payload delivery",
    "filename": "Artifacts dropped",
}
