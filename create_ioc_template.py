import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "IOC Intelligence"

# ── Column Definitions ──
columns = [
    ("indicator",           20, "The IOC value (IP, domain, hash, URL)"),
    ("ioc_type",            15, "ip-dst / domain / url / md5 / sha256 / sha1 / email"),
    ("threat_actor",        20, "e.g. APT28, Lazarus, FIN7"),
    ("threat_type",         22, "Phishing / Ransomware / APT etc."),
    ("confidence",          15, "Low / Medium / High / Critical"),
    ("first_seen",          15, "YYYY-MM-DD"),
    ("last_seen",           15, "YYYY-MM-DD"),
    ("country_origin",      18, "Country of threat origin (e.g. RU, CN, KP)"),
    ("targeted_country",    18, "Country being targeted (e.g. IN, US, UA)"),
    ("targeted_region",     18, "e.g. South Asia, Europe, Middle East"),
    ("targeted_sector",     18, "e.g. Finance, Healthcare, Government"),
    ("associated_malware",  20, "e.g. Emotet, Cobalt Strike, Ryuk"),
    ("associated_cve",      18, "e.g. CVE-2021-44228"),
    ("tags",                25, "Comma-separated MISP tags"),
    ("source",              25, "e.g. VirusTotal, AlienVault, Manual Research"),
    ("reference_url",       35, "Source article/report URL"),
    ("notes",               35, "Any additional analyst notes"),
]

# ── Header Style ──
header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
header_font = Font(bold=True, color="00d4ff", size=11)
border = Border(
    bottom=Side(style='medium', color='00d4ff')
)

for col_idx, (col_name, width, comment) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 35

# ── Data Validations ──
# IOC Type dropdown
dv_ioc = DataValidation(
    type="list",
    formula1='"ip-dst,ip-src,domain,url,md5,sha256,sha1,email-src,filename"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Invalid IOC Type",
    error="Please select a valid IOC type from the list"
)
ws.add_data_validation(dv_ioc)
dv_ioc.sqref = "B2:B1000"

# Threat Type dropdown
dv_threat = DataValidation(
    type="list",
    formula1='"Phishing,Ransomware,APT/Espionage,C2 Infrastructure,Malware Distribution,DDoS,Supply Chain,Vulnerability Exploitation"',
    allow_blank=True
)
ws.add_data_validation(dv_threat)
dv_threat.sqref = "D2:D1000"

# Confidence dropdown
dv_conf = DataValidation(
    type="list",
    formula1='"Low,Medium,High,Critical"',
    allow_blank=True
)
ws.add_data_validation(dv_conf)
dv_conf.sqref = "E2:E1000"

# ── Sample Data Rows ──
sample_data = [
    ["185.220.101.45", "ip-dst", "Lazarus Group", "C2 Infrastructure", "High",
     "2024-01-15", "2024-06-01", "KP", "IN", "South Asia",
     "Finance", "Cobalt Strike", "CVE-2021-44228",
     "tlp:amber,misp-galaxy:threat-actor=\"Lazarus Group\"",
     "Manual Research", "https://example.com/report1",
     "Active C2 observed in Lazarus campaign targeting Indian banks"],

    ["evil-phishing.com", "domain", "APT28", "Phishing", "Medium",
     "2024-03-01", "2024-05-15", "RU", "US,EU", "Europe",
     "Government", "Fancy Bear Implant", "",
     "tlp:white,misp-galaxy:threat-actor=\"APT28\"",
     "AlienVault OTX", "https://example.com/report2",
     "Phishing domain mimicking government login portal"],

    ["44d88612fea8a8f36de82e1278abb02f", "md5", "FIN7", "Ransomware", "Critical",
     "2024-04-10", "2024-06-05", "UA", "IN", "South Asia",
     "Healthcare", "BlackCat", "CVE-2023-4966",
     "tlp:red,misp-galaxy:threat-actor=\"FIN7\"",
     "VirusTotal", "https://example.com/report3",
     "Ransomware hash targeting hospital infrastructure"],
]

# ── Row Styles ──
row_fill_1 = PatternFill(start_color="0d1117", end_color="0d1117", fill_type="solid")
row_fill_2 = PatternFill(start_color="161b22", end_color="161b22", fill_type="solid")
row_font = Font(color="c9d1d9", size=10)

for row_idx, row_data in enumerate(sample_data, start=2):
    fill = row_fill_1 if row_idx % 2 == 0 else row_fill_2
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.font = row_font
        cell.alignment = Alignment(vertical='center', wrap_text=False)

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"

# ── Instructions Sheet ──
ws2 = wb.create_sheet("Instructions")
ws2.sheet_properties.tabColor = "00d4ff"
instructions = [
    ["MISP IOC Ingestion Template — Instructions"],
    [""],
    ["COLUMN GUIDE:"],
    ["indicator", "The actual IOC value — IP, domain, hash, URL, email"],
    ["ioc_type", "MISP attribute type: ip-dst, domain, url, md5, sha256, sha1, email-src"],
    ["threat_actor", "Name of APT/threat group (e.g. APT28, Lazarus, FIN7)"],
    ["threat_type", "Select from dropdown: Phishing, Ransomware, APT/Espionage, C2 etc."],
    ["confidence", "Select from dropdown: Low / Medium / High / Critical"],
    ["first_seen", "Date first observed — format YYYY-MM-DD"],
    ["last_seen", "Date last observed — format YYYY-MM-DD"],
    ["country_origin", "2-letter ISO country code of threat origin (RU, CN, KP, IR)"],
    ["targeted_country", "2-letter ISO code of targeted country (IN, US, UA)"],
    ["targeted_region", "Region being targeted (South Asia, Europe, Middle East)"],
    ["targeted_sector", "Industry sector (Finance, Healthcare, Government, Energy)"],
    ["associated_malware", "Malware family name (Emotet, Cobalt Strike, Ryuk)"],
    ["associated_cve", "CVE ID if applicable (CVE-2021-44228)"],
    ["tags", "Comma-separated MISP tags (tlp:amber, misp-galaxy:threat-actor=...)"],
    ["source", "Where you found this IOC (VirusTotal, Manual Research, AlienVault)"],
    ["reference_url", "Full URL of source article or report"],
    ["notes", "Your analyst notes about this IOC"],
    [""],
    ["WORKFLOW:"],
    ["1. Fill in IOC Intelligence sheet with your research"],
    ["2. Save the file as ioc_intel.xlsx"],
    ["3. Run: python3 misp_ingest.py"],
    ["4. IOCs will be automatically pushed to MISP with all tags"],
    ["5. Verify in MISP UI: https://localhost/events/index"],
]

for row in instructions:
    ws2.append(row)

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 60

output = '/home/mrutyu1998/ioc_intel.xlsx'
wb.save(output)
print(f"Template saved: {output}")
